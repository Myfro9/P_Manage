import pandas as pd
import numpy as np
import os
import datetime
from datetime import timedelta
import time
import xlwings as xw
import shutil
import numbers
import Contract_gen
import rd_DataBase
import openpyxl
import openpyxl.worksheet
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers, is_date_format
import re



#TimeStr = time.strftime("%Y-%m-%d", time.localtime(time.time()))
TimeStr = datetime.datetime.now().strftime("%Y-%m-%d-%H_%M")
TimeStr_14days = (datetime.datetime.now()+ datetime.timedelta(days=14)).strftime("%Y-%m-%d")
TimeStr_21days = (datetime.datetime.now()+ datetime.timedelta(days=21)).strftime("%Y-%m-%d")

'''产生生产计划安排，包括：
    1）领料清单，以及供应商摧料清单
    2）外加工委外清单、领料清单、供应商催料清单
    3）
    
'''


#### 指定生产计划单号
#Task = pd.DataFrame({'Task':['RW202106-C1','YF202106-A'] }) # 可以包含子生产计划单
'''pd_Task = pd.DataFrame({'Task':['RW202106-B','RW202106-D','RW202106-E','RW202106-F','RW202106-H','RW202107-A','RW202107-B',
                                'RW202107-C','RW202107-D','RW202108-A','RW202108-B','RW202108-D','RW202109-A','RW202109-B'] })'''  # 可以包含子生产计划单
#pd_Task = pd.DataFrame({'Task':['RW202201-A-1','RW202201-A-2','RW202201-A-3','RW202201-A-4','RW202201-A-5'] })

#pd_Task = pd.DataFrame({'Task':['RW202201-A-1','RW202201-A-2','RW202201-A-3','RW202201-A-4','RW202201-A-5','RW202201-A-6','RW202201-A-7','RW202201-A-8']})
#pd_Task = pd.DataFrame({'Task':['RW202201-A-1','RW202201-A-2'] })
# 国网清频项目
pd_Task_HH = pd.DataFrame({'Task':['RW202112-G-1','RW202112-G-2','RW202112-H-1','RW202112-H-2','RW202112-H-3',\
                                'RW202201-A-4','RW202201-A-6','RW202201-A-7','RW202201-A-8','RW202201-A-9',\
                                'RW202201-B-1','RW202201-B-2','RW202201-B-3','RW202201-B-4',\
                                'RW202201-C-1','RW202201-C-2','RW202201-C-3','RW202201-C-4','RW202201-C-5',\
                                'RW202201-C-6','RW202201-C-7','RW202201-D','RW202204-A-5',\
                                'RW202203-V-1','RW202203-V-2','RW202203-V-3','RW202203-V-4','RW202203-V-5'] })
'''pd_Task = pd.DataFrame({'Task':[ 'RW202112-G-1','RW202112-G-2','RW202201-A-5'
                               ] })'''

# 美国项目
File_USTask_NameStr = './Purchase_Rawdata/00在产任务单/美国在产生产任务单-2022.5.25.xlsx'
pd_Task0 = pd.DataFrame(pd.read_excel(File_USTask_NameStr))
pd_Task_US = pd_Task0.loc[:,['任务单号']]

pd_Task_US.columns = ['Task']


pd_Task = pd_Task_US

pd_VirtualStock,pd_PrjList,FileNameStr_PrjList,StockInfor_Filename,FileNameStr_contract,\
           Contract_FileNameStr1,Contract_FileNameStr2,FileNameStr_outstock,FileNameStr_instock,\
           FileNameStr_outsourcing,FileNameStr_outsourcingBack,\
            pd_StockInfor,pd_Contract_inTask,pd_Outstock_inTask,pd_Instock_inTask,pd_OutSourcing_inTask,pd_OutSourcingBack_inTask = rd_DataBase.Purchase_Rawdata_analyze(None)

#### 指定采购询价结果（张永亮），（其中包含了合同、付款信息）
FolderNameStr_quote = './Purchase_Rawdata/23询价结果/'
FileNameStr_quote = '合并汇总表-版本87-20220608.xlsx'
#FileNameStr_quote1 = '汇总表2-版本11-20210928.xlsx'
pd_QuoteInfo0 = pd.DataFrame(pd.read_excel(FolderNameStr_quote+FileNameStr_quote,sheet_name='操作'))
pd_QuoteInfo = pd_QuoteInfo0

''''#######  导入在产生产任务单
FileNameStr_PrjList = './Purchase_Rawdata/00在产任务单/近期在产生产任务单-2022-good.xls'
pd_PrjList = pd.DataFrame(pd.read_excel(FileNameStr_PrjList))'''

#### 导入采购到货记录（张永亮）  --- I （注意，不是采购入库记录，到货与入库之间有时间差！），后面可以滤出在产任务单的到货记录
FolderNameStr = './Purchase_Rawdata/24采购到货记录/'
FileNameStr = '2021ERP到货单列表-20220210.xls'   # 2021 年记录
FileNameStr_1 = '2022ERP到货单列表-20220610.xls' # 2022 年记录
pd_PurchaseArrive_0 = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr))
pd_PurchaseArrive_1 = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr_1))
pd_PurchaseArrive = pd.concat([pd_PurchaseArrive_1,pd_PurchaseArrive_0])
pd_PurchaseArrive_inTask = rd_DataBase.OnGoing_Prjinfo_Filter(pd_PrjList, pd_PurchaseArrive, 0)
####  导入ERP材料入库记录（许金鲍） -- II ，并整理出对应在产任务单的材料入库记录
'''FolderNameStr = './Purchase_Rawdata/12材料入库记录/'
FileNameStr_instock2021 = '采购入库单列表 2012.31.xls' # 2021年记录
FileNameStr_instock2022 = '采购入库单列表2022.1.31.xls'
pd_ERPInStock2022 = pd.DataFrame(pd.read_excel(FolderNameStr+FileNameStr_instock2022))
pd_ERPInStock2021 = pd.DataFrame(pd.read_excel(FolderNameStr+FileNameStr_instock2021))
pd_ERPInStock = pd.concat([pd_ERPInStock2021,pd_ERPInStock2022])
pd_ERPInstock_inTask = rd_DataBase.OnGoing_Prjinfo_Filter(pd_PrjList, pd_ERPInStock,0)'''
pd_ERPInstock_inTask = pd_Instock_inTask
####  导入ERP材料出库记录（许金鲍） -- III ，并整理出对应在产任务单的材料出库记录
'''FolderNameStr = './Purchase_Rawdata/13材料出库记录/'
FileNameStr_outstock2021 = '材料出库单列表2021.12.31.xls' # 2021年记录
FileNameStr_outstock2022 = '材料出库单2022.1.31.xls'
pd_ERPOutStock2021 = pd.DataFrame(pd.read_excel(FolderNameStr+FileNameStr_outstock2021))
pd_ERPOutStock2022 = pd.DataFrame(pd.read_excel(FolderNameStr+FileNameStr_outstock2022))
pd_ERPOutStock = pd.concat([pd_ERPOutStock2021,pd_ERPOutStock2022])
pd_ERPOutstock_inTask = rd_DataBase.OnGoing_Prjinfo_Filter(pd_PrjList, pd_ERPOutStock,0)'''
pd_ERPOutstock_inTask = pd_Outstock_inTask
####  导入委外材料出库记录（许金鲍） --- IV
'''FolderNameStr = './Purchase_Rawdata/15委外材料出库记录/'
FileNameStr_ERPOutsourcing_info2021 = '委外材料出库单列表2021.12.31.xls'
FileNameStr_ERPOutsourcing_info2022 = '委外材料出库列表2022.1.31.xls'
pd_ERPOutsourcing_info2021 = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr_ERPOutsourcing_info2021))
pd_ERPOutsourcing_info2022 = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr_ERPOutsourcing_info2022))
pd_ERPOutsourcing= pd.concat([pd_ERPOutsourcing_info2021,pd_ERPOutsourcing_info2022])
pd_ERPOutsourcing_inTask = rd_DataBase.OnGoing_Prjinfo_Filter(pd_PrjList, pd_ERPOutsourcing,0)  # 宽松过滤，采用.contains()过滤'''
pd_ERPOutsourcing_inTask = pd_OutSourcing_inTask
#### 导入ERP库存信息（许金鲍），将相同ERP编码的库存项进行合并  --- V
'''StockInfor_Filename2021 = './Purchase_Rawdata/11库存记录/erp现存量列表2021.12.31.XLS'
StockInfor_Filename2022 = './Purchase_Rawdata/11库存记录/erp现存量2022.1.31.XLS'
pd_StockInfor2021 = pd.DataFrame(pd.read_excel(StockInfor_Filename2021))
pd_StockInfor2022 = pd.DataFrame(pd.read_excel(StockInfor_Filename2022))
pd_StockInfor = pd.concat([pd_StockInfor2021,pd_StockInfor2022])'''
# 直接使用rd_DataBase.Purchase_Rawdata_analyze()函数的输出结果

#### 导入仓库点料记录信息（何伟） -- VI
FolderNameStr = './Purchase_Rawdata/31生产点料记录/'
FileNameStr = '2022年点料信息5.13.xls'
pd_Outstock_info = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr))



#### 导入采购追料备注信息，分海外（张琳）和国内（王莹莹）
FolderNameStr = './Purchase_Rawdata/25采购追料备注/'
FileNameStr = '海华物料ERP到货明细表20220309V1.xls' # 国内（王莹莹）
pd_Purchase_Trackinfor = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr))


####### 导入供应商信息（张永亮）
FolderNameStr = './Purchase_Rawdata/22供应商档案/'
FileNameStr = '供应商档案 57-20220524.xlsx'
pd_Supplier_Infor = pd.DataFrame(pd.read_excel(FolderNameStr + FileNameStr))

def PruchaseTable_PurchaseArrive_allTask_gen(pd_Task,pd_QuoteInfo,pd_PurchaseArrive_inTask,pd_Outstock_info):
    log_FileNameStr = './PurchaseBOM_询价表_log.xlsx'
    pd_log = pd.DataFrame(pd.read_excel(log_FileNameStr))
    key_list = ['RW', 'YF', 'GC']
    cnt0 =0
    for i in range(pd_Task.shape[0]):
        # 读取PurchaseBOM_询价表_log.xlsx, 并统计所有任务单的某个ERP编码的总需求数量，新采购数量
        task = pd_Task.loc[pd_Task.index[i],'任务单号']

        PurchaseBOM_str0 = pd_log.loc[pd_log['生产计划单']==task,'PurchaseBOM']
        if not PurchaseBOM_str0.empty:
            PurchaseBOM_str = PurchaseBOM_str0.values[0]
            PurchaseTable_filestr = './Results/' + PurchaseBOM_str + '/' + 'PurchaseTable' + PurchaseBOM_str.strip('PurchaseBOM_') + '.xls'
            pd_PurchaseTable = pd.DataFrame(pd.read_excel(PurchaseTable_filestr))
            pd_PurchaseTable = pd_PurchaseTable[pd_PurchaseTable['生产计划单号']==task]
            pd_PurchaseTable_combined = pd_PurchaseTable.groupby('ERP').sum()
            pd_PurchaseTable_combined['ERP'] = pd_PurchaseTable_combined.index
            pd_PurchaseTable_combined['生产计划单'] = task
            pd_PurchaseTable_combined.index = range(0,pd_PurchaseTable_combined.shape[0])

            if cnt0 ==0:
                pd_PurchaseTable_allTask = pd_PurchaseTable_combined
            else:
                pd_PurchaseTable_allTask = pd.concat([pd_PurchaseTable_allTask,pd_PurchaseTable_combined])

            # 根据张永亮提供的采购合并汇总表pd_QuoteInfo，以及2021年和2022年的ERP到货单列表pd_PurchaseArrive_inTask，统计所有生产计划单的某个ERP编码对应合同的到货情况（包括合并采购、特殊采购、全波另外库存等情况）

            pd_QuoteInfo_Task0 = pd_QuoteInfo[pd_QuoteInfo['生产计划单号'] == task].dropna(subset=['渠道'])
            if not pd_QuoteInfo_Task0.empty:
                pd_QuoteInfo_Task0['报价数量'].fillna(0,inplace=True)
                pd_QuoteInfo_Task0['单价（元）'].fillna(0,inplace=True)
                pd_QuoteInfo_grouped = pd_QuoteInfo_Task0.groupby('ERP编码')
                cnt = 0
                for item, item_df in pd_QuoteInfo_grouped:
                    pd_QuoteInfo_grouped2 = item_df.groupby('渠道').sum()
                    pd_QuoteInfo_grouped2['渠道'] = pd_QuoteInfo_grouped2.index
                    pd_QuoteInfo_grouped2['ERP编码'] = item
                    pd_QuoteInfo_grouped2['型号'] = item_df['型号'].values[0]
                    pd_QuoteInfo_grouped2.index = range(0,pd_QuoteInfo_grouped2.shape[0])
                    if cnt ==0:
                        pd_QuoteInfo_Task = pd_QuoteInfo_grouped2
                    else:
                        pd_QuoteInfo_Task = pd.concat([pd_QuoteInfo_Task,pd_QuoteInfo_grouped2])
                    cnt = cnt +1
                #if task == 'RW202105-B':
                    #print(task)
                pd_QuoteInfo_Task['生产计划单号'] = task
                pd_QuoteInfo_Task['合同到货量'] = 0
                pd_QuoteInfo_Task['合并合同备注'] = None
                pd_QuoteInfo_Task.index = range(0, pd_QuoteInfo_Task.shape[0])
                pd_QuoteInfo_Task = pd_QuoteInfo_Task[pd_QuoteInfo_Task['生产计划单号'] != pd_QuoteInfo_Task['渠道']]
                pd_QuoteInfo_Task.index = range(0, pd_QuoteInfo_Task.shape[0])
                for j in range(pd_QuoteInfo_Task.shape[0]):
                    ERP = pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j],'ERP编码']
                    Supplier = pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j],'渠道']
                    #print(Supplier)
                    if any(key in Supplier for key in key_list)  :   ## 针对当前的采购任务包含在其它任务单里的情况，即供应商信息为其它的生产任务单号
                        task2 = Supplier
                        result1 = pd_QuoteInfo[(pd_QuoteInfo['生产计划单号'] == task2) & (pd_QuoteInfo['ERP编码'] == ERP)]  # 查找另外的生产计划单所对应的采购信息
                        if not result1.empty:  # 在合并采购中找到了相对应的采购信息
                            temp = pd_PurchaseArrive_inTask.loc[\
                                (pd_PurchaseArrive_inTask['存货编码(cinvcode)']==ERP) & (pd_PurchaseArrive_inTask['备注(cmemo)'].str.contains(task2) ),'主计量数量(iquantity)' ]
                            if not temp.empty:
                                pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合同到货量'] = temp.sum()
                            pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合并合同备注'] = task2
                    elif '全波' in Supplier:
                        pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合并合同备注'] = '全波有其它库存'
                    elif '特殊采购' in Supplier:  ## 特殊采购的情况
                        result1 = pd_QuoteInfo[(pd_QuoteInfo['渠道'] == '特殊采购') & (pd_QuoteInfo['ERP编码'] == ERP)]
                        if not  result1.empty:
                            pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合并合同备注'] = '特殊采购'
                            pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合同到货量'] = pd_PurchaseArrive_inTask.loc[\
                                (pd_PurchaseArrive_inTask['存货编码(cinvcode)'] == ERP) & (\
                                    pd_PurchaseArrive_inTask['备注(cmemo)'].str.contains('特殊采购')), '主计量数量(iquantity)'].sum()
                    else: ## 正常情况，即当前任务单的询价信息里可以查到对应的供应商信息
                        temp = pd_PurchaseArrive_inTask.loc[ \
                            (pd_PurchaseArrive_inTask['存货编码(cinvcode)'] == ERP) & ( \
                                pd_PurchaseArrive_inTask['备注(cmemo)'].str.contains(task)), '主计量数量(iquantity)']
                        if not temp.empty:
                            pd_QuoteInfo_Task.loc[pd_QuoteInfo_Task.index[j], '合同到货量'] = temp.sum()
                if cnt0 ==0:
                    pd_QuoteInfo_Track = pd_QuoteInfo_Task[['生产计划单号','ERP编码','型号','需求数量','报价数量','渠道','合同到货量','合并合同备注']]
                else:
                    pd_QuoteInfo_Track = pd.concat([pd_QuoteInfo_Track,pd_QuoteInfo_Task[['生产计划单号','ERP编码','型号','需求数量','报价数量','渠道','合同到货量','合并合同备注']]])
                cnt0 = cnt0 + 1

    pd_PurchaseTable_allTask.index = range(0,pd_PurchaseTable_allTask.shape[0])
    pd_QuoteInfo_Track.index = range(0,pd_QuoteInfo_Track.shape[0])
    pd_PurchaseTable_allTask_grouped = pd_PurchaseTable_allTask.groupby('ERP')
    cnt =0
    for item, item_df in pd_PurchaseTable_allTask_grouped:
        #对于每一个ERP，统计各个任务单的总需求，以及各个任务单的需求明细
        ERP = item
        Total_Num = item_df['任务单总需求量'].sum()
        for ii in range(item_df.shape[0]):
            msg_new = '{' + item_df.loc[item_df.index[ii], '生产计划单'] + '需求' + str(
                item_df.loc[item_df.index[ii], '任务单总需求量']) + '}\n'
            if ii ==0 :
                msg = msg_new
            else:
                msg = msg + msg_new
        # 对于每一个ERP，统计各个任务单的总到货数，以及各个任务单的到货明细
        pd_QuoteInfo_Track_ERP = pd_QuoteInfo_Track[pd_QuoteInfo_Track['ERP编码']==ERP]
        Total_Arrive_Num_ERP = pd_QuoteInfo_Track_ERP['合同到货量'].sum()
        if not pd_QuoteInfo_Track_ERP.empty:
            for i in range(pd_QuoteInfo_Track_ERP.shape[0]):
                msg_arrive_newline = '{' + pd_QuoteInfo_Track_ERP.loc[pd_QuoteInfo_Track_ERP.index[i],'生产计划单号'] + '到货' \
                                     + str(pd_QuoteInfo_Track_ERP.loc[pd_QuoteInfo_Track_ERP.index[i],'合同到货量']) + '_合并采购备注：_' +\
                                     str(pd_QuoteInfo_Track_ERP.loc[pd_QuoteInfo_Track_ERP.index[i],'合并合同备注'])+ '}\n'
                if i ==0:
                    msg_arrive = msg_arrive_newline
                else:
                    msg_arrive = msg_arrive + msg_arrive_newline
        else:
            msg_arrive = ''

        # 对于每一个ERP，统计各个任务单的总点料数，以及各个任务单的点料明细
        pd_Outstock_ERP = pd_Outstock_info[pd_Outstock_info['子件编码']==ERP]
        Total_Outstock_Num_ERP = pd_Outstock_ERP['送料数量'].fillna(0).sum()
        if not pd_Outstock_ERP.empty:
            for i in range(pd_Outstock_ERP.shape[0]):
                msg_outstock_newline = '{' + pd_Outstock_ERP.loc[pd_Outstock_ERP.index[i],'生产计划单号'] + '点料' \
                + str(pd_Outstock_ERP.loc[pd_Outstock_ERP.index[i],'送料数量']) + '}\n'
                if i ==0:
                    msg_outstock = msg_outstock_newline
                else:
                    msg_outstock = msg_outstock + msg_outstock_newline
        else:
            msg_outstock = ''
        # 新建一个dataFrame
        pd_Newline = pd.DataFrame({'ERP': [ERP], '多任务总需求量': [Total_Num], '多任务总需求明细': [msg],\
                                   '多任务总到货量':[Total_Arrive_Num_ERP],'多任务总到货明细': [msg_arrive],\
                                   '多任务点料总数量':[Total_Outstock_Num_ERP],'多任务点料明细':[msg_outstock]})
        if cnt ==0:

            pd_PurchaseTable_analyse = pd_Newline
        else:
            pd_PurchaseTable_analyse = pd.concat([pd_PurchaseTable_analyse,pd_Newline])
        cnt = cnt +1
    #if ERP == 'M0101004009':
        #print('test1')

#print('test')
    return pd_PurchaseTable_analyse


## 搜索目录

def Current_Stock_gen(pd_ERPInStock,pd_PurchaseArrive,pd_ERPOutStock,pd_Outstock,pd_StockInfor,pd_ERPOutsourcing_info):
    # pd_PurchaseArrive 相同ERP，生产任务单的项合并
    pd_PurchaseArrive_grouped = pd_PurchaseArrive.groupby('备注(cmemo)')
    i =0
    for item, item_df in pd_PurchaseArrive_grouped:
        pd_PurchaseArrive_grouped2 = item_df.groupby('存货编码(cinvcode)').sum()
        pd_PurchaseArrive_grouped2['生产计划单'] = item
        pd_PurchaseArrive_grouped2['ERP'] = pd_PurchaseArrive_grouped2.index
        pd_PurchaseArrive_grouped2.index = range(0,pd_PurchaseArrive_grouped2.shape[0])
        if i == 0:
            pd_PurchaseArrive_combined = pd_PurchaseArrive_grouped2
        else:
            pd_PurchaseArrive_combined = pd.concat([pd_PurchaseArrive_combined,pd_PurchaseArrive_grouped2])
        i = i+1
    pd_PurchaseArrive_combined.index = range(0,pd_PurchaseArrive_combined.shape[0])

    # pd_ERPInstock 相同ERP、生产任务单的项合并
    pd_ERPInStock_grouped = pd_ERPInStock.groupby('备注(cMemo)')
    i =0
    for item1,item_df1 in pd_ERPInStock_grouped:
        pd_ERPInStock_grouped2 = item_df1.groupby('存货编码(cInvCode)').sum()
        pd_ERPInStock_grouped2['生产计划单号'] = item1
        pd_ERPInStock_grouped2['ERP'] = pd_ERPInStock_grouped2.index
        pd_ERPInStock_grouped2.index = range(0,pd_ERPInStock_grouped2.shape[0])
        if i ==0:
            pd_ERPInStock_combined = pd_ERPInStock_grouped2
        else:
            pd_ERPInStock_combined = pd.concat([pd_ERPInStock_combined,pd_ERPInStock_grouped2])
        i =i+1
    pd_ERPInStock_combined.index = range(0,pd_ERPInStock_combined.shape[0])

    # 在pd_PurchaseArrive 中按生产计划单和ERP逐条搜索pd_ERPInstock, 并记录ERP入库数量，计算出已经到货，但没有入库的数量 Num1
    pd_PurchaseArrive_combined['已入库数量'] =0
    for i in range(pd_PurchaseArrive_combined.shape[0]):
        ERP = pd_PurchaseArrive_combined.loc[pd_PurchaseArrive_combined.index[i],'ERP']
        Task = pd_PurchaseArrive_combined.loc[pd_PurchaseArrive_combined.index[i],'生产计划单']
        result = pd_ERPInStock_combined[(pd_ERPInStock_combined['ERP'] == ERP) & (pd_ERPInStock_combined['生产计划单号'] == Task)]
        if not result.empty:
            pd_PurchaseArrive_combined.loc[pd_PurchaseArrive_combined.index[i],'已入库数量'] = result['数量(iQuantity)'].values[0]
    pd_PurchaseArrive_combined['已到货未入库数量'] = pd_PurchaseArrive_combined.apply(lambda x: x['主计量数量(iquantity)'] - x['已入库数量'], axis=1)
    pd_PurchaseArrive_combined.loc[pd_PurchaseArrive_combined['已到货未入库数量'] < 0, '已到货未入库数量'] = 0
    pd_PurchaseArrive_NotinERP = pd_PurchaseArrive_combined[pd_PurchaseArrive_combined['已到货未入库数量'] > 0]

    # pd_Outstock （何伟）相同ERP，生产任务单的项合并
    pd_Outstock_grouped = pd_Outstock.groupby('生产计划单号')
    i = 0
    for item, item_df in pd_Outstock_grouped:
        pd_Outstock_grouped2 = item_df.groupby('子件编码').sum()
        pd_Outstock_grouped2['生产计划单'] = item
        pd_Outstock_grouped2['ERP'] = pd_Outstock_grouped2.index
        pd_Outstock_grouped2.index = range(0, pd_Outstock_grouped2.shape[0])
        if i == 0:
            pd_Outstock_combined = pd_Outstock_grouped2
        else:
            pd_Outstock_combined = pd.concat([pd_Outstock_combined, pd_Outstock_grouped2])
        i = i + 1
    pd_Outstock_combined.index = range(0, pd_Outstock_combined.shape[0])
    pd_Outstock_combined['送料数量']=pd_Outstock_combined['送料数量'].fillna(0)

    # pd_ERPOutstock 相同ERP，生产任务单的项合并
    pd_ERPOutStock_grouped = pd_ERPOutStock.groupby('备注(cMemo)')
    i = 0
    for item1, item_df1 in pd_ERPOutStock_grouped:
        pd_ERPOutStock_grouped2 = item_df1.groupby('材料编码(cInvCode)').sum()
        pd_ERPOutStock_grouped2['生产计划单号'] = item1
        pd_ERPOutStock_grouped2['ERP'] = pd_ERPOutStock_grouped2.index
        pd_ERPOutStock_grouped2.index = range(0, pd_ERPOutStock_grouped2.shape[0])
        if i == 0:
            pd_ERPOutStock_combined = pd_ERPOutStock_grouped2
        else:
            pd_ERPOutStock_combined = pd.concat([pd_ERPOutStock_combined, pd_ERPOutStock_grouped2])
        i = i + 1
    pd_ERPOutStock_combined.index = range(0, pd_ERPOutStock_combined.shape[0])

    # pd_ERPOutsourcing_info 相同ERP，生产计划单的项合并
    pd_ERPOutsourcing_grouped = pd_ERPOutsourcing_info.groupby('备注(cMemo)')
    i = 0
    for item1, item_df1 in pd_ERPOutsourcing_grouped:
        pd_ERPOutsourcing_grouped = item_df1.groupby('材料编码(cInvCode)').sum()
        pd_ERPOutsourcing_grouped['生产计划单号'] = item1
        pd_ERPOutsourcing_grouped['ERP'] = pd_ERPOutsourcing_grouped.index
        pd_ERPOutsourcing_grouped.index = range(0, pd_ERPOutsourcing_grouped.shape[0])
        if i == 0:
            pd_ERPOutsourcing_combined = pd_ERPOutsourcing_grouped
        else:
            pd_ERPOutsourcing_combined = pd.concat([pd_ERPOutsourcing_combined, pd_ERPOutsourcing_grouped])
        i = i + 1
    pd_ERPOutsourcing_combined.index = range(0, pd_ERPOutsourcing_combined.shape[0])
    pd_ERPOutStock_combined = pd.concat([pd_ERPOutsourcing_combined,pd_ERPOutStock_combined])
    pd_ERPOutStock_combined.index = range(0, pd_ERPOutStock_combined.shape[0])

    # 为了计算当前库存，首先对 pd_PurchaseArrive_combined， pd_ERPInStock_combined，
    # 在pd_Outstock 中按生产计划单和ERP逐条搜索pd_ERPOutStock，并记录ERP出库数量，计算出已经点料，但还没有ERP出库记录的数量 Num2
    pd_Outstock_combined['已出库数量'] = 0
    for i in range(pd_Outstock_combined.shape[0]):
        ERP = pd_Outstock_combined.loc[pd_Outstock_combined.index[i], 'ERP']
        Task = pd_Outstock_combined.loc[pd_Outstock_combined.index[i], '生产计划单']
        result = pd_ERPOutStock_combined[
            (pd_ERPOutStock_combined['ERP'] == ERP) & (pd_ERPOutStock_combined['生产计划单号'] == Task)]
        if not result.empty:
            pd_Outstock_combined.loc[pd_Outstock_combined.index[i], '已出库数量'] = \
            result['数量(iQuantity)'].values[0]
    pd_Outstock_combined['已点料未出库数量'] = pd_Outstock_combined.apply(lambda x: x['送料数量'] - x['已出库数量'],
                                                                           axis=1)
    pd_Outstock_combined.loc[pd_Outstock_combined['已点料未出库数量']<0, '已点料未出库数量'] =0
    pd_Outstock_NotinERP = pd_Outstock_combined[pd_Outstock_combined['已点料未出库数量']>0]

    # 在pd_StockInfor 中，先对ERP编码进行合并。
    pd_StockInfor_grouped = pd_StockInfor.groupby('存货编码').sum()
    pd_StockInfor_grouped['存货编码'] = pd_StockInfor_grouped.index
    pd_StockInfor_grouped.index = range(0,pd_StockInfor_grouped.shape[0])

    # 再对pd_StockInfor 按照ERP编码，加入Num1 ，再减掉Num2
    for i in range(pd_PurchaseArrive_NotinERP.shape[0]):
        ERP = pd_PurchaseArrive_NotinERP.loc[pd_PurchaseArrive_NotinERP.index[i],'ERP']
        Num1 = pd_PurchaseArrive_NotinERP.loc[pd_PurchaseArrive_NotinERP.index[i],'已到货未入库数量']
        #if ERP == 'M0102000357':
        #    pd_test = pd_StockInfor.loc[pd_StockInfor['存货编码'] == ERP,'现存数量']
        #    print('test1')
        result = pd_StockInfor_grouped.loc[pd_StockInfor_grouped['存货编码'] == ERP, '现存数量']
        if not result.empty:
            pd_StockInfor_grouped.loc[pd_StockInfor_grouped['存货编码'] == ERP,'现存数量'] = result + Num1
        else:
            pd_New_item = pd.DataFrame({'现存数量': [Num1], '存货编码': [ERP]})
            pd_StockInfor_grouped = pd.concat([pd_StockInfor_grouped,pd_New_item])

    for i in range(pd_Outstock_combined.shape[0]):
        ERP = pd_Outstock_combined.loc[pd_Outstock_combined.index[i],'ERP']
        Num2 = pd_Outstock_combined.loc[pd_Outstock_combined.index[i],'已点料未出库数量']
        result = pd_StockInfor_grouped.loc[pd_StockInfor_grouped['存货编码'] == ERP,'现存数量']
        if not result.empty:
            pd_StockInfor_grouped.loc[pd_StockInfor_grouped['存货编码'] == ERP,'现存数量'] = result - Num2
        else:
            print('Warning: ERP库存里没有 ', ERP,'， 但点料记录中却有该料的记录，且还没有出库？' )


    return pd_StockInfor_grouped


def PurchaseBOM_proc(Task,pd_BOMfile, pd_QuoteInfo,pd_Supplier_Infor,pd_PurchaseArrive,pd_Purchase_Trackinfor,pd_Current_Stock,pd_Outstock_info,pd_PurchaseTable_analyse):

    pd_BOMfile['供应商'] = None
    pd_BOMfile['交货方式'] = None
    pd_BOMfile['合同签订时间'] = None
    pd_BOMfile['预付款支付时间'] = None
    pd_BOMfile['计划交期'] = pd.to_datetime('')
    pd_BOMfile['到货数量'] = 0
    pd_BOMfile['flag'] = None   # 内部标签列，用于标注状态，用于openpyxl后期修改excel文件
    pd_BOMfile['到货跟踪信息'] = None
    pd_BOMfile['当前库存量'] = 0
    pd_BOMfile['已点料数量'] =0
    pd_BOMfile['多任务总需求量'] =0
    pd_BOMfile['多任务总需明细'] = None
    pd_BOMfile['多任务总到货量'] =0
    pd_BOMfile['多任务总到货明细'] =None
    pd_BOMfile['多任务总点料量'] = 0
    pd_BOMfile['多任务总点料明细'] =None
    key_list = ['RW', 'YF', 'GC']
    key_list1 = ['M0301', 'M0302', 'M0304','M0306'] ### 结构件BOM
    #print([x for x in pd_BOMfile.columns[pd_BOMfile.columns.str.contains('母件编码')]][0])
    # print(pd_BOMfile.columns[pd_BOMfile.columns.str.contains('母件编码')].values)
    # pd_BOMfile.columns[pd_BOMfile.columns.str.contains('母件编码')].values = ['母件编码'] # 对于带有特殊字符的Column，进行更换
    #pd_BOMfile = pd_BOMfile.rename(columns=lambda x: x.replace("*", "").replace('"', '')).replace(" ", "") # 去除columns中的特殊字符和空格
    pd_BOMfile = pd_BOMfile.rename(columns=lambda x: x.replace("'", "").replace('"', '')).replace(" ", "")
    pd_Outstock_info['子件编码'].str.replace('\'', '')
    pd_Outstock_info['母件编码'].str.replace('\'', '')
    ##pd_PurchaseTable_analyse = PruchaseTable_PurchaseArrive_allTask_gen(pd_Task, pd_QuoteInfo, pd_PurchaseArrive,pd_Outstock_info)
    for i in range(pd_BOMfile.shape[0]):
        Parent_ERP = pd_BOMfile.loc[pd_BOMfile.index[i],[x for x in pd_BOMfile.columns[pd_BOMfile.columns.str.contains('母件编码')]][0]]
        ERP = pd_BOMfile.loc[pd_BOMfile.index[i], '子件编码(cpscode)']
        Num_instock = pd_Current_Stock.loc[pd_Current_Stock['存货编码'] == ERP,'现存数量']
        pd_PurchaseTrack = pd_PurchaseTable_analyse[pd_PurchaseTable_analyse['ERP'] == ERP]
        if not pd_PurchaseTrack.empty:
            pd_BOMfile.loc[pd_BOMfile.index[i],'多任务总需求量'] = pd_PurchaseTrack.loc[:,'多任务总需求量'].values[0]
            pd_BOMfile.loc[pd_BOMfile.index[i], '多任务总需明细'] = pd_PurchaseTrack.loc[:, '多任务总需求明细'].values[0]
            pd_BOMfile.loc[pd_BOMfile.index[i], '多任务总到货量'] = pd_PurchaseTrack.loc[:, '多任务总到货量'].values[0]
            pd_BOMfile.loc[pd_BOMfile.index[i], '多任务总到货明细'] = pd_PurchaseTrack.loc[:, '多任务总到货明细'].values[0]
            pd_BOMfile.loc[pd_BOMfile.index[i], '多任务总点料量'] = pd_PurchaseTrack.loc[:,  '多任务点料总数量'].values[0]
            pd_BOMfile.loc[pd_BOMfile.index[i], '多任务总点料明细'] = pd_PurchaseTrack.loc[:, '多任务点料明细'].values[0]

        if 'RW202106-C' in Task:  # 对 'RW202106-C' 特别对待，zyl的手工错误，导致RW202106-C-1～12 被手工合并到一起了，这只是一个特例
            Task0 = 'RW202106-C'
        else:
            Task0 = Task
        if not Num_instock.empty:
            pd_BOMfile.loc[pd_BOMfile.index[i],'当前库存量'] = Num_instock.values[0]

        #if (Parent_ERP == 'M0106002108') and (ERP == 'M0102002773') and (Task0 == 'RW202201-A-1'):
        #    print('test1:')
        Num_outstock = pd_Outstock_info.loc[
            (pd_Outstock_info['生产计划单号'] == Task0) & (pd_Outstock_info['子件编码'] == ERP) & (pd_Outstock_info[
                '母件编码'] == Parent_ERP),'送料数量']
        if not Num_outstock.empty:
            pd_BOMfile.loc[pd_BOMfile.index[i],'已点料数量'] = Num_outstock.values[0]
        if pd_BOMfile.loc[pd_BOMfile.index[i],'需新增采购量'] >0 :

            Type = pd_BOMfile.loc[pd_BOMfile.index[i], '子件名称(cinvname)']
            Component = pd_BOMfile.loc[pd_BOMfile.index[i], '是否为半成品']

            result = pd_QuoteInfo[(pd_QuoteInfo['生产计划单号'] == Task0) & (pd_QuoteInfo['ERP编码'] == ERP) ]
            if not result.empty:
                #prePay_infor = result['预付款'].fillna(0).values[-1]
                supplier = str(result['渠道'].values[-1])
                if any(key in supplier for key in key_list):   ## 针对当前的采购任务包含在其它任务单里的情况，即供应商信息为其它的生产任务单号
                    result1 = pd_QuoteInfo[(pd_QuoteInfo['生产计划单号'] == supplier) & (pd_QuoteInfo['ERP编码'] == ERP)]  # 查找另外的生产计划单所对应的采购信息
                    if not result1.empty:   # 在合并采购中找到了相对应的采购信息
                        supplier1 = str(result1['渠道'].values[-1])
                        sr_prePay_infor = pd_Supplier_Infor.loc[pd_Supplier_Infor['供应商名称'] == supplier1, '账期']
                        if sr_prePay_infor.shape[0] >0 :
                            pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = supplier1
                            prePay_infor = str(sr_prePay_infor.values[-1])
                        else:
                            print('Wanning # 1: 供应商信息中，供应商信息中，没有{}的信息'.format(supplier1))
                            pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = ''
                            prePay_infor = ''

                        ##  合同签订时间
                        Contract_time_date = result1['盖章日期'].dt.date.values[-1]
                        Contract_time = result1['盖章日期'].values[-1]
                        if np.isnat(Contract_time):
                            print('Warning # 4-1: {}，的合并采购合同没有签订日期'.format(ERP), '供应商是：', supplier1)
                            pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 1  # 合同没有签订日期
                        pd_BOMfile.loc[pd_BOMfile.index[i], '合同签订时间'] = Contract_time_date

                        ##  到货情况
                        result2 = pd_PurchaseArrive[(pd_PurchaseArrive['备注(cmemo)'] == supplier) & (pd_PurchaseArrive['存货编码(cinvcode)'] == ERP)]
                        if not result2.empty:
                            ArriveNUM = result2['主计量数量(iquantity)'].values.sum()
                            supplier_Arrive = result2['供应商(cvenabbname)']
                            pd_BOMfile.loc[pd_BOMfile.index[i], '到货数量'] = ArriveNUM
                        ## 采购跟踪备注
                        result3 = pd_Purchase_Trackinfor[(pd_Purchase_Trackinfor['合同编号'] == supplier) & (pd_Purchase_Trackinfor['存货编号(cInvCode)'] == ERP)]
                        if not result3.empty:
                            Trackinfo = pd_Purchase_Trackinfor['备注'].values[0]
                            pd_BOMfile.loc[pd_BOMfile.index[i],'到货跟踪信息'] = Trackinfo

                    else:   # 合并采购信息中也没有找到相对应的采购合同
                        print('Warning # 2: ',  supplier, '中也没有', ERP, '的采购信息','该物料是：', Type )
                        pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = ''
                        prePay_infor = ''
                elif '全波' in supplier:   ## 针对供应商为'全波'的情况
                    pd_BOMfile.loc[pd_BOMfile.index[i],'供应商'] = '全波有其它库存'
                    prePay_infor = '_'
                elif '特殊采购' in supplier: ## 特殊采购的情况
                    pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = '全波有特别采购'
                    prePay_infor = '_'
                    ## 特殊采购的到货情况
                    result2 = pd_PurchaseArrive[(pd_PurchaseArrive['备注(cmemo)'] == supplier) & (pd_PurchaseArrive['存货编码(cinvcode)'] == ERP)]
                    if not result2.empty:
                        ArriveNUM = result2['主计量数量(iquantity)'].values
                        supplier_Arrive = result2['供应商(cvenabbname)']
                        pd_BOMfile.loc[pd_BOMfile.index[i], '到货数量'] = ArriveNUM

                else:                     ## 正常情况，即当前任务单的询价信息里可以查到对应的供应商信息
                    sr_prePay_infor = pd_Supplier_Infor.loc[pd_Supplier_Infor['供应商名称'] == supplier, '账期']
                    if Component != 'Yes' and not '焊接' in Type and str(supplier) != 'nan': # 排除半成品、焊接，以及供应商栏信息为空的情况
                        if not sr_prePay_infor.shape[0] >0:
                            print('Waring # 3: 供应商信息中，没有{}的信息'.format(supplier))
                            pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = '还未完成采购询价'
                            prePay_infor = ''
                        else:
                            prePay_infor = str(sr_prePay_infor.values[-1])
                            pd_BOMfile.loc[pd_BOMfile.index[i], '供应商'] = supplier
                            result3 = pd_Purchase_Trackinfor[(pd_Purchase_Trackinfor['合同编号'] == Task0) & (
                                        pd_Purchase_Trackinfor['存货编号(cInvCode)'] == ERP)]
                            if not result3.empty:
                                Trackinfo = pd_Purchase_Trackinfor['备注'].values[0]
                                pd_BOMfile.loc[pd_BOMfile.index[i], '到货跟踪信息'] = Trackinfo
                    else:
                        prePay_infor = ''   ## 半成品或焊接

                    ##  合同签订时间

                    Contract_time_date = result['盖章日期'].dt.date.values[-1]
                    Contract_time = result['盖章日期'].values[-1]
                    if np.isnat(Contract_time):
                        print('Warning # 4-2: {}，的合同没有签订日期'.format(ERP), '供应商是：', supplier)
                        pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 1  # 合同没有签订日期
                    pd_BOMfile.loc[pd_BOMfile.index[i], '合同签订时间'] = Contract_time_date

                    ## 到货信息
                    result2 = pd_PurchaseArrive[(pd_PurchaseArrive['备注(cmemo)'] == Task0) & (pd_PurchaseArrive['存货编码(cinvcode)'] == ERP)]
                    if not result2.empty:
                        if result2.shape[0]>1:
                            ArriveNUM = result2['主计量数量(iquantity)'].sum()
                        else:
                            ArriveNUM = result2['主计量数量(iquantity)'].values
                        supplier_Arrive = result2['供应商(cvenabbname)']
                        #print(Task)
                        #print(ERP)
                        #if (Task == 'RW202107-B-1') & (ERP == 'M0101023006'):
                            #print('test')
                        pd_BOMfile.loc[pd_BOMfile.index[i], '到货数量'] = ArriveNUM

                #根据付款信息，推算计划到货时间信息
                #print(prePay_infor)
                if '预付' in prePay_infor and '付清' in prePay_infor:
                    if any(key in ERP for key in key_list1):  ## 判断不同的产品种类，交货周期会不同
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '付款后3周，全款提货'
                        pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 21
                        #payTime = pd_BOMfile.loc[pd_BOMfile.index[i], '预付款支付时间']
                        #payTime = '2021-9-30'
                        #str_payTime = pd.to_datetime(payTime)
                        #print(str_payTime)
                        #pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = str_payTime
                    else:
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '付款后2周，全款提货'
                        pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 14
                elif '预付' in prePay_infor:
                    if any(key in ERP for key in key_list1):
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '付款后3周'
                        pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 21
                    else:
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '付款后2周'
                        pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 14
                elif '款到发货' in prePay_infor:
                    if any(key in ERP for key in key_list1):
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '合同签订3周后，全款提货'
                        Contract_time1 = pd.to_datetime(Contract_time_date)
                        pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = (Contract_time1 + datetime.timedelta(days=14))
                    else:
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '合同签订2周后，全款提货'
                        Contract_time1 = pd.to_datetime(Contract_time_date)
                        #print(Contract_time1)
                        pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = (Contract_time1 + datetime.timedelta(days=14))
                elif prePay_infor == '_':
                        Contract_time1 = pd.to_datetime('')
                        #print(Contract_time1)
                        pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = Contract_time1
                else:
                    if any(key in ERP for key in key_list1):
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '合同签订后3周'
                        Contract_time1 = pd.to_datetime(Contract_time_date)
                        #print(Contract_time1)
                        pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = (Contract_time1 + datetime.timedelta(days=21))
                    else:
                        pd_BOMfile.loc[pd_BOMfile.index[i], '交货方式'] = '合同签订后2周'
                        Contract_time1 = pd.to_datetime(Contract_time_date)
                        #print(Contract_time1)
                        pd_BOMfile.loc[pd_BOMfile.index[i], '计划交期'] = (Contract_time1 + datetime.timedelta(days=14))



            else:
                if Component != 'Yes' and not '焊接' in Type:
                    print('Warning # 5: 需要采购物料{}，但是没有找到对应的采购合同！'.format(ERP),'该物料是：', Type )
                    pd_BOMfile.loc[pd_BOMfile.index[i], 'flag'] = 0  # 没有找到合同

        # 补充'多任务总需求量'，'多任务总需明细'


    pd_BOMfile['预付款支付时间'] = pd.to_datetime(pd_BOMfile['预付款支付时间'])
    pd_BOMfile['预付款支付时间'] = pd_BOMfile['预付款支付时间'].dt.date
    pd_BOMfile['计划交期'] = pd_BOMfile['计划交期'].dt.date


    #print('test')

    return pd_BOMfile

def OutsourcingOrder_gen(pd_BOMfile,root_lower,Dst_folder):
    for i in range(pd_BOMfile.shape[0]):
        ERP = pd_BOMfile.loc[pd_BOMfile.index[i],'子件编码(cpscode)']
        if pd_BOMfile.loc[pd_BOMfile.index[i],'是否为半成品'] == 'Yes' and pd_BOMfile.loc[pd_BOMfile.index[i],'需新增采购量'] >0:
            if os.path.exists(root_lower+'/'+ERP +'.xls'):
                if os.path.exists(Dst_folder+'/'+ERP +'.xls'):
                    os.remove(Dst_folder+'/'+ERP +'.xls')
                shutil.copy(root_lower+'/'+ERP +'.xls', Dst_folder+'/'+ERP +'.xls')
            else:
                print('Warning # 5: 在',root_lower , '中缺少半成品{}的BOM文件'.format(ERP))


def openpyxl_modify(BOM_file_result):
    wb = openpyxl.load_workbook(BOM_file_result)
    ws = wb.active
    N_row = ws.max_row
    fill_red = PatternFill("solid", fgColor="FFFF0000")  # fgColor="1874CD"
    fill_blue = PatternFill("solid", fgColor="1874CD")
    fill_yellow = PatternFill("solid", fgColor="00FFFF00")
    fill_orrange = PatternFill("solid", fgColor="00FF6600")
    for i in range(2, N_row+1):
        flag = ws.cell(row=i, column=21-2).value
        paymentTime = ws.cell(row=i, column=18-2)
        deliverType = ws.cell(row=i,column=16-2).value
        deliverTime = ws.cell(row=i, column=19-2)
        supplier = ws.cell(row=i,column=15-2)
        contractTime = ws.cell(row=i,column=17-2)
        if flag != None:
            if int(flag) > 4:
                #print('the flag is:', flag)
                paymentTime.fill = fill_red
                #print(deliverType)
                if '付款' in deliverType:
                    deliverTime.value = "=P" + str(i) + "+" + str(int(flag))
                elif '合同' in deliverType:
                    deliverTime.value = "=O" + str(i) + "+" + str(int(flag))
                else:
                    deliverTime.value =''
                deliverTime.number_format = numbers.FORMAT_DATE_YYYYMMDD2
            elif int(flag) ==0: # 没有对应的采购合同
                supplier.fill = fill_blue
            elif int(flag) ==1: # 没有合同签订日期
                contractTime.fill = fill_yellow


    ws.delete_cols(19,1)  # 删除flag 内部标签列
    wb.save(BOM_file_result)




def find_BOMlevel(BOM_Folder):
    L = 1
    for root, dirs, files in os.walk(BOM_Folder):
        if root.split('/')[-1]=='L5':
            L = 5
        elif root.split('/')[-1]=='L4':
            if L<5:
                L = 4
        elif root.split('/')[-1] == 'L3':
            if L<4:
                L = 3
        elif root.split('/')[-1] == 'L2':
            if L<3:
                L =2
        else:
            if L<2:
                L =1
    return L

'''def SearchPurchaseBOM(Task):
    pd_PurchaseBOM_log = pd.DataFrame(pd.read_excel('./PurchaseBOM_询价表_log.xlsx'))
    PurchaseBOM = pd_PurchaseBOM_log[pd_PurchaseBOM_log['生产计划单'] == Task]
    assert not PurchaseBOM.empty, 'PurchaseBOM 记录中没有{}的记录'.format(Task)
    FolderNameStr_PurchaseBOM = './Results/' + PurchaseBOM +'/'
    return FolderNameStr_PurchaseBOM '''

###############################################################################
############# Main Function  #################################################
###############################################################################

#pd_subTask = pd.DataFrame(columns=['生产任务单号','母件编码','母件名称','层级','子件编码','子件名称','到货日期','委外加工/内部派单日期'])
pd_subTask = pd.DataFrame(columns=['生产任务单号','母件编码','母件名称','层级','子件编码','子件名称','到货情况'])
Folder_schedule = './Results/Schedule/'+TimeStr +'/'

if not os.path.exists(Folder_schedule):
    os.makedirs(Folder_schedule)

Current_Stock = Current_Stock_gen(pd_ERPInstock_inTask,pd_PurchaseArrive_inTask,pd_ERPOutstock_inTask,pd_Outstock_info,pd_StockInfor,pd_ERPOutsourcing_inTask)
pd_PurchaseTable_analyse = PruchaseTable_PurchaseArrive_allTask_gen(pd_PrjList, pd_QuoteInfo, pd_PurchaseArrive,pd_Outstock_info)

for i in range(pd_Task.shape[0]):
    Kwd = str(pd_Task.iloc[i, 0])  #注意Kwd不是生产计划单号
    #### 指定到货安排结果存放地址
    Folder_result = Folder_schedule + Kwd + '_到货安排_' + TimeStr + '/'
    if not os.path.exists(Folder_result):
        os.makedirs(Folder_result)

    #### 指定委外安排结果存放地址
    #Folder_result1 = './Results/Schedule/' + Kwd + '_委外安排_' + TimeStr + '/'
    #if not os.path.exists(Folder_result1):
        #os.makedirs(Folder_result1)


    pd_QuoteInfo = Contract_gen.Clean_Quote_info(pd_QuoteInfo,pd_Supplier_Infor)
    Err = Contract_gen.Check_quote_info(pd_QuoteInfo,pd_Supplier_Infor)

    # 将缺料PurchaseBOM文件夹中对应该生产计划单的PurchaseBOM文件夹拷贝到目标文件夹中
    pd_PurchaseBOM_log = pd.DataFrame(pd.read_excel('./PurchaseBOM_询价表_log.xlsx'))
    Target_tasklist = pd_PurchaseBOM_log[pd_PurchaseBOM_log['生产计划单'].str.contains(Kwd)]
    if not Target_tasklist.empty:
        for ii in range(Target_tasklist.shape[0]):
            Target_task = Target_tasklist.loc[Target_tasklist.index[ii], '生产计划单']
            PurchaseBOM = pd_PurchaseBOM_log[pd_PurchaseBOM_log['生产计划单'] == Target_task]['PurchaseBOM'].values[0]
            FolderNameStr_PurchaseBOM = './Results/' + PurchaseBOM +'/'
            if os.path.exists(FolderNameStr_PurchaseBOM):
                for root, dirs, files in os.walk(FolderNameStr_PurchaseBOM):
                    if Target_task in root.split('/')[-1]:
                        #print('The root:' ,root)
                        if os.path.exists(Folder_result+root.split('/')[-1]):
                            shutil.rmtree(Folder_result+root.split('/')[-1] )
                        shutil.copytree(root,Folder_result+root.split('/')[-1] )
    else:
        assert False,'PurchaseBOM 记录中没有{}的记录'.format(Kwd)

    for root, dirs, files in os.walk(Folder_result):
        for file in files:
            if file.split('.')[-1].lower() == 'xls' or file.split('.')[-1].lower() == 'xlsx':   # 搜索BOM文件夹下L1、L2、L3、L4下的每一个子目录，对其中每一个BOM文件，补充供应商到料信息
                #print(root+'/'+file)

                BOM_file = root+'/'+file.lower()
                BOM_file_result = '.'+BOM_file.split('.')[-2] + '.xlsx'
                pd_BOMfile = pd.DataFrame(pd.read_excel(BOM_file))
                pd_BOMfile = pd_BOMfile[['母件编码 *(cpspcode)H','母件名称 *(cinvname)H','子件编码(cpscode)','子件名称(cinvname)','规格型号(cinvstd)','主计量单位(ccomunitname)',
                                         '基本用量分子(ipsquantity)','基本用量分母(tdqtyd)','任务单总需求量','需新增采购量','是否为半成品']]
                os.remove(BOM_file)
                # 在BOM清单中加入供应商到货信息
                Task = BOM_file.split('/')[-3]  #Task 是生产计划单号


                pd_BOMfile = PurchaseBOM_proc(Task,pd_BOMfile, pd_QuoteInfo,pd_Supplier_Infor,pd_PurchaseArrive_inTask,pd_Purchase_Trackinfor,Current_Stock,pd_Outstock_info,pd_PurchaseTable_analyse)
                pd_BOMfile.to_excel(BOM_file_result)

                # 对BOM_file 进行二次格式修改
                #print(BOM_file_result)
                #if BOM_file_result == './Results/Schedule/RW202108-B_到货安排_2021-09-28/RW202108-B-4/L1/p0402000364.xlsx':
                    #print('test')
                #openpyxl_modify(BOM_file_result)


    # 增加委外任务到货信息
    for root0, dirs0, files0 in os.walk(Folder_result):
        if ((root0 != Folder_result) &  ('20' in root0.split('/')[-1])):
            Folder_task = root0
            task = Folder_task.split('/')[-1]
            print(Folder_task)
            L = find_BOMlevel(Folder_task)
            #print(L)
            for i in range(1,L):
                k = L-i #倒序搜索目录
                #print(k)
                for root, dirs, files in os.walk(Folder_task + '/L' + str(k) +'/'): # 搜索倒数第二层BOM目录
                    for file in files:
                        if file.split('.')[-1].lower() == 'xls' or file.split('.')[-1].lower() == 'xlsx':
                            BOM_file1 = root + '/' + file.lower()
                            pd_BOMfile1 = pd.DataFrame(pd.read_excel(BOM_file1))
                            if 'Unnamed: 0' in pd_BOMfile1.columns :
                                pd_BOMfile1 = pd_BOMfile1.drop(labels=['Unnamed: 0'],axis=1)
                            for j in range(pd_BOMfile1.shape[0]):
                                if (pd_BOMfile1.loc[pd_BOMfile1.index[j], '是否为半成品'] == 'Yes')& (pd_BOMfile1.loc[pd_BOMfile1.index[j],'需新增采购量']>0):
                                    ERP_Lowlevel = pd_BOMfile1.loc[pd_BOMfile1.index[j], '子件编码(cpscode)']
                                    #if ERP_Lowlevel == 'M0106000419':
                                    #    print('Debug')
                                    ready = 1
                                    for root_L2, dirs_L2, files_L2 in os.walk(Folder_result+ task + '/L' + str(k+1) +'/'): # 搜索下层BOM目录
                                        for file_L2 in files_L2:
                                            if (file_L2.split('.')[-1].lower() == 'xls' or file_L2.split('.')[-1].lower() == 'xlsx') & (file_L2.split('.')[-2] == ERP_Lowlevel.lower()):
                                                BOM_file_Lowlevel = root_L2 + '/' + file_L2.lower()
                                                pd_BOMfile_Lowlevel = pd.DataFrame(pd.read_excel(BOM_file_Lowlevel))
                                                for jj in range(pd_BOMfile_Lowlevel.shape[0]):
                                                    if not ((pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '到货数量'] >= pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '需新增采购量'])
                                                        or (pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '是否为半成品'] == '半成品部件已到货')
                                                        or ('H' in pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '子件编码(cpscode)'])
                                                        or (pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '供应商']== '全波有其它库存')
                                                        or (pd_BOMfile_Lowlevel.loc[pd_BOMfile_Lowlevel.index[jj], '供应商']== '特殊采购')):
                                                        ready =0
                                    if ready ==1:
                                        pd_BOMfile1.loc[pd_BOMfile1.index[j], '是否为半成品'] = '半成品部件已到货'
                            pd_BOMfile1.to_excel(BOM_file1)

    # 颜色注释等
    for root, dirs, files in os.walk(Folder_result):
        for file in files:
            if file.split('.')[-1].lower() == 'xls' or file.split('.')[-1].lower() == 'xlsx':
                BOM_file = root + '/' + file.lower()
                openpyxl_modify(BOM_file)

    # 生成各个生产任务单的子任务单列表
    for root, dirs, files in os.walk(Folder_result):
        for file in files:
            if file.split('.')[-1].lower() == 'xls' or file.split('.')[-1].lower() == 'xlsx':
                level = root.split('/')[-1]
                task1 = root.split('/')[-2]
                BOM_file2 = root + '/' + file.lower()
                pd_BOMfile2 = pd.DataFrame(pd.read_excel(BOM_file2))
                for i in range(pd_BOMfile2.shape[0]):
                    if (pd_BOMfile2.loc[pd_BOMfile2.index[i], '是否为半成品'] == 'Yes') or (pd_BOMfile2.loc[pd_BOMfile2.index[i], '是否为半成品'] == '半成品部件已到货'):
                        ERP_parent = pd_BOMfile2.loc[pd_BOMfile2.index[i], '母件编码 *(cpspcode)H']
                        Name_parent = pd_BOMfile2.loc[pd_BOMfile2.index[i], '母件名称 *(cinvname)H']
                        ERP_child = pd_BOMfile2.loc[pd_BOMfile2.index[i], '子件编码(cpscode)']
                        Name_child = pd_BOMfile2.loc[pd_BOMfile2.index[i], '子件名称(cinvname)']
                        if (pd_BOMfile2.loc[pd_BOMfile2.index[i], '是否为半成品'] == '半成品部件已到货'):
                            Purchase_arrive = '已到货'
                        elif pd_BOMfile2.loc[pd_BOMfile2.index[i], '需新增采购量'] == 0 :
                            Purchase_arrive = '有库存'
                        else:
                            Purchase_arrive = ''
                        pd_subTask0 = pd.DataFrame({'生产任务单号':[task1], '母件编码':[ ERP_parent], '母件名称': [Name_parent], '层级': [level],'子件编码':[ ERP_child],'子件名称':[ Name_child],'到货情况':[Purchase_arrive]})
                        if pd_subTask.empty:
                            pd_subTask = pd_subTask0
                        else:
                            pd_subTask = pd.concat([pd_subTask,pd_subTask0])




pd_subTask['到货日期'] = None
pd_subTask['到货日期'] = pd.to_datetime(pd_subTask['到货日期'])
pd_subTask['到货日期'] = pd_subTask['到货日期'].dt.date
pd_subTask['委外加工/内部派单日期'] = None
pd_subTask['委外加工/内部派单日期'] = pd.to_datetime(pd_subTask['委外加工/内部派单日期'])
pd_subTask['委外加工/内部派单日期'] = pd_subTask['委外加工/内部派单日期'].dt.date
pd_subTask.to_excel(Folder_schedule +'生产派单任务列表_' + TimeStr +'.xlsx')


'''pd_subTasks = pd.DataFrame(columns=['ERP','L2','L3','L4'])
    indx = 0
    #print(Folder_result+Kwd)
    for root, dirs, files in os.walk(Folder_result+Kwd+ '/L1/'):  # 搜索L1子目录
        for file in files:
            if file.split('.')[-1].lower() == 'xls' or file.split('.')[-1].lower() == 'xlsx':
                BOM_file_L1 = root + '/' + file.lower()
                pd_BOMfile_L1 = pd.DataFrame(pd.read_excel(BOM_file_L1))
                for i in range(pd_BOMfile_L1.shape[0]):
                    ready_L1 = 1
                    if (pd_BOMfile_L1.loc[pd_BOMfile_L1.index[i],'是否为半成品'] == 'Yes') & (pd_BOMfile_L1.loc[pd_BOMfile_L1.index[i],'需新增采购量']>0):
                        ERP_L2 = pd_BOMfile_L1.loc[pd_BOMfile_L1.index[i],'子件编码(cpscode)']
                        ready_L2 = 1
                        if pd_subTasks.empty:
                            pd_subTasks = pd.DataFrame(index=[0], columns=['ERP','L2', 'L3', 'L4'])
                            pd_subTasks.iloc[0, 1] = (file.split('.')[-2])
                            pd_subTasks.iloc[0, 0] = ERP_L2
                        else:
                            pd_subTasks.iloc[indx, 1] = (file.split('.')[-2])
                            pd_subTasks.iloc[indx, 0] = ERP_L2
                        indx = indx + 1

                        for root_L2, dirs_L2, files_L2 in os.walk(Folder_result+Kwd+ '/L2/'):  ##在L2中搜索相应的ERP文件
                            for file_L2 in files_L2:
                                if (file_L2.split('.')[-1].lower() == 'xls' or file_L2.split('.')[-1].lower() == 'xlsx') & file_L2.split('.')[-2] == ERP_L2:
                                    BOM_file_L2 = root_L2 + '/' + file_L2.lower()
                                    pd_BOMfile_L2 = pd.DataFrame(pd.read_excel(BOM_file_L2))
                                    for j in range(pd_BOMfile_L2.shape[0]):
                                        if (pd_BOMfile_L2.loc[pd_BOMfile_L2.index[i], '是否为半成品'] == 'Yes') & (
                                                pd_BOMfile_L2.loc[pd_BOMfile_L2.index[i], '需新增采购量'] > 0):
                                            ERP_L3 = pd_BOMfile_L2.loc[pd_BOMfile_L2.index[i], '子件编码(cpscode)']
                                            ready_L3 = 1
                                            pd_subTasks.iloc[indx, 2] = (file_L2.split('.')[-2])  # L3
                                            pd_subTasks.iloc[indx, 0] = ERP_L2     # L3
                                            indx = indx +1

                                            for root_L3, dirs_L3, files_L3 in os.walk(
                                                Folder_result + Kwd + '/L3/'):  ##在L2中搜索相应的ERP文件
                                                for file_L3 in files_L3:
                                                    if (file_L3.split('.')[-1].lower() == 'xls' or file_L3.split('.')[
                                                        -1].lower() == 'xlsx') & file_L3.split('.')[-2] == ERP_L3:
                                                        BOM_file_L3 = root_L3 + '/' + file_L3.lower()
                                                        pd_BOMfile_L3 = pd.DataFrame(pd.read_excel(BOM_file_L3))
                                                        for k in range(pd_BOMfile_L3.shape[0]):
                                                            if (pd_BOMfile_L3.loc[
                                                                    pd_BOMfile_L3.index[i], '是否为半成品'] == 'Yes') & (
                                                                    pd_BOMfile_L3.loc[
                                                                        pd_BOMfile_L3.index[i], '需新增采购量'] > 0):
                                                                ERP_L4 = pd_BOMfile_L3.loc[
                                                                    pd_BOMfile_L3.index[i], '子件编码(cpscode)']
                                                                ready_L4 = 1
                                                                pd_subTasks.iloc[indx, 3] = (file_L3.split('.')[-2])  # L4
                                                                pd_subTasks.iloc[indx, 0] = ERP_L3  # L4
                                                                indx = indx + 1






                if root.split('/')[-1] == 'L1':
                root_lower = Folder_result  + Task + '/L2'
                if os.path.exists(root_lower):
                    OutsourcingOrder_gen(pd_BOMfile,root_lower,Folder_result1)
            elif root.split('/')[-1] == 'L2':
                root_lower = Folder_result +  Task + '/L3'
                if os.path.exists(root_lower):
                    OutsourcingOrder_gen(pd_BOMfile, root_lower, Folder_result1)
            elif root.split('/')[-1] == 'L3':
                root_lower = Folder_result +  Task + '/L4'
                if os.path.exists(root_lower):
                    OutsourcingOrder_gen(pd_BOMfile, root_lower, Folder_result1)'''




